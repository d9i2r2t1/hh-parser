import logging
import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from hh_parser.lib import HhParser

log = logging.getLogger(__name__)


class ReportFileProcessor:
    """Обработка файла с результами парсинга."""

    def __init__(self, hh_parsed_data: HhParser.HhParserResults) -> None:
        self.__report_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'temp')
        self.__formats = {'background_col_name': PatternFill('solid', fgColor='D6DCE3')}
        self.__parsing_results = hh_parsed_data.df_current_jobs
        self.__search_text = hh_parsed_data.search_text

    parsing_results = property(lambda self: self.__parsing_results)

    def _get_report_file_name(self) -> str:
        """Получи название файла с результатами парсинга."""
        return f'{self.__search_text}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'.replace(' ', '_')

    def create_report_file(self) -> str:
        """
        Создай файл с результатами парсинга.
        :return: Путь к созданному файлу
        """
        file_path = os.path.join(self.__report_folder, self._get_report_file_name())
        self.__parsing_results.to_excel(file_path, index=False)
        self._format_file(file_path)
        log.info(f'Report file created: {file_path}')
        return file_path

    def _format_file(self, file_path: str) -> None:
        """
        "Причеши" файл с результатами парсинга.
        :param file_path: Путь к файлу
        """
        wb = load_workbook(filename=file_path)
        ws = wb['Sheet1']

        for col in range(ws.max_column):
            # Выделяем фон ячеек с названием столбцов
            for cell in list(ws.columns)[col]:
                if cell.row == 1:
                    cell.fill = self.__formats['background_col_name']

            # Выставляем ширину столбцов
            length = max([len(str(cell.value)) for cell in list(ws.columns)[col]])
            ws.column_dimensions[get_column_letter(col + 1)].width = length

        wb.save(filename=file_path)
        log.debug('Report file formatted')
