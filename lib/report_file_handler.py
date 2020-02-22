import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

log = logging.getLogger(__name__)


class ReportFileHandler:
    """Обработка файла с результами парсинга"""

    def __init__(self, file_path):
        self.file_path = file_path
        self._formats = {'background_col_name': PatternFill('solid', fgColor='D6DCE3')}

    def format_file(self):
        """Форматирование .xlsx-файла"""

        wb = load_workbook(filename=self.file_path)
        ws = wb['Sheet1']

        for col in range(ws.max_column):
            # Выделяем фон ячеек с названием столбцов
            for cell in list(ws.columns)[col]:
                if cell.row == 1:
                    cell.fill = self._formats['background_col_name']

            # Выставляем ширину столбцов
            length = max([len(str(cell.value)) for cell in list(ws.columns)[col]])
            ws.column_dimensions[get_column_letter(col + 1)].width = length

        wb.save(filename=self.file_path)
        log.info('Report file formatted')
